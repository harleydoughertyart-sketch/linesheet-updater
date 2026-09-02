"""Read the ranked-sales workbook — the source of truth for line sheet figures.

Only three fields matter to a line sheet: the style's ``Row#``, its ``Qty Sold``,
and its ``$Sales``. Everything else in the workbook is working-out.

Two traps this module exists to avoid:

- ``Row#`` is *not* ``Rank by Qty Sold``. They diverge on 38 of 151 rows in the
  Women's sheet. The line sheet prints ``Row#``.
- A style appears in several sheets with a *different* ``Row#`` in each
  (``LX1015`` is 38 in Women's, 2 in Luxe, 30 in Sweaters). The caller must name
  the sheet; there is no sensible default.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROW_NUM = "Row#"
STYLE = "Style #"
DESCRIPTION = "Style Description"
QTY = "Qty Sold"
SALES = "$Sales"

#: The ranking column is renamed between seasons — Fall 2026 heads it "Row#",
#: Spring 2027 "Numb" — while every other column keeps its name. Match on any
#: known spelling rather than failing on a workbook that is otherwise identical.
ROW_NUM_ALIASES = (ROW_NUM, "Numb", "Row #", "Row", "No.", "No", "#")


def _row_num_column(columns: dict[str, int],
                    headers: list[str] | None = None) -> int | None:
    """Which column holds the ranking the line sheet prints.

    Sheets in the *same workbook* disagree about whether that column has a
    heading at all. In the July 2026 report "Fall 2026 Accessories" heads it
    ``Row#``, while "Fall 2026" and every fabric breakdown beneath it leave the
    header cell empty — the column is there, first, holding 1, 2, 3…, just
    unlabelled. Rejecting those sheets threw away every Women's style in the
    book and left the picker to fall back on Accessories, which reported all 143
    styles as missing.

    So an unheaded *first* column is accepted as the ranking. It is a narrow
    allowance: only column A, only when its header is blank, and every caller
    still requires the value to be a whole number, which is what keeps the
    trailing ``SubTtl`` row and any stray text column out.
    """
    for name in ROW_NUM_ALIASES:
        if name in columns:
            return columns[name]
    if headers and not headers[0]:
        return 0
    return None


@dataclass(frozen=True)
class SalesRow:
    """What the line sheet needs to know about one style.

    An empty cell reads as zero, deliberately: the report is regenerated each
    season, so a style with no figure in it has not sold, and the sheet should
    say so rather than keep last season's number. See :func:`_whole`.
    """

    style: str
    row_num: int
    qty: int
    sales: float


@dataclass(frozen=True)
class Window:
    """A few workbook rows around one style, for showing it in situ."""

    headers: list[str]
    rows: list[list]
    #: index into ``rows`` of the style asked for, or None if it isn't here
    highlight: int | None
    #: 1-based row number in the worksheet, so the reviewer can find it
    sheet_row: int | None


@dataclass(frozen=True)
class Grid:
    """The worksheet as it appears, not as the parser interprets it.

    The review card shows the figures where they actually live — with their
    neighbours and their real worksheet row number — so a reviewer can check the
    app against the file rather than taking its word.
    """

    headers: list[str]
    rows: list[list]
    #: style number -> index into ``rows``
    index: dict[str, int]
    #: index into ``rows`` -> 1-based worksheet row
    sheet_rows: list[int]
    #: which column the ranking was found in — not always the one called ``Row#``
    row_at: int = 0

    def window(
        self, style: str, context: int = 2, columns: list[str] | None = None
    ) -> Window:
        """Rows around ``style``, narrowed to the columns worth showing.

        The real workbook has 46 columns, most of them running totals and
        per-rep breakdowns. Showing all of them would bury the four a reviewer
        actually needs to check.
        """
        wanted = columns or [ROW_NUM, STYLE, DESCRIPTION, QTY, SALES]
        picks = [self.headers.index(name) for name in wanted if name in self.headers]
        # The ranking column is the one figure the reviewer is checking, so it
        # has to appear even on the sheets that leave its header blank.
        if columns is None and self.row_at not in picks:
            picks.insert(0, self.row_at)
        shown = [self.headers[i] or ROW_NUM for i in picks]
        position = self.index.get(style)
        if position is None:
            return Window(shown, [], None, None)
        start = max(0, position - context)
        stop = min(len(self.rows), position + context + 1)
        return Window(
            headers=shown,
            rows=[[row[i] for i in picks] for row in self.rows[start:stop]],
            highlight=position - start,
            sheet_row=self.sheet_rows[position],
        )


def _is_row_number(value: object) -> bool:
    """Whether a cell holds a ranking rather than a label or a stray flag.

    ``bool`` is a subclass of ``int`` in Python, so a TRUE in this column would
    otherwise pass as row 1 and a FALSE as row 0 — a plausible ranking invented
    out of a checkbox.
    """
    return isinstance(value, int) and not isinstance(value, bool)


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _whole(value: object) -> int:
    """A quantity. An empty or unreadable cell counts as zero.

    The client's rule, and it is the right one for this document: the report is
    generated per season, so a style with no figure in it has not sold, and the
    line sheet should say so rather than leave last season's number standing.
    Zero is a fact here, not a guess.
    """
    if _blank(value) or isinstance(value, bool):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _money(value: object) -> float:
    """A sales figure. An empty or unreadable cell counts as zero — see
    :func:`_whole`.

    Accepts the text a hand-edited cell can hold — "$12,578.40" — because a
    workbook that has been through someone's hands is still a workbook.
    """
    if _blank(value) or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    try:
        amount = float(text)
    except ValueError:
        return 0.0
    return -amount if negative else amount


def _norm(value: object) -> str:
    """Collapse whitespace in a header cell.

    Header cells carry embedded newlines (``"Qty\\nSold"``), so a literal
    comparison against ``"Qty Sold"`` silently misses.
    """
    return re.sub(r"\s+", " ", str(value)).strip()


def read_grid(xlsx_path: str | Path, sheet_name: str) -> Grid:
    """The worksheet's raw cells, indexed by style number.

    Unlike :func:`read_sales` this keeps every column and the real worksheet row
    numbers, because the point is to show the reviewer the file itself.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"no sheet named {sheet_name!r}; workbook has {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise KeyError(f"sheet {sheet_name!r} is empty") from None

        headers = [_norm(cell) if cell is not None else "" for cell in header]
        columns = {name: index for index, name in enumerate(headers) if name}
        style_at = columns.get(STYLE)
        row_at = _row_num_column(columns, headers)
        if style_at is None or row_at is None:
            raise KeyError(f"sheet {sheet_name!r} lacks {STYLE!r} or {ROW_NUM!r}")

        kept: list[list] = []
        sheet_rows: list[int] = []
        index: dict[str, int] = {}
        for offset, row in enumerate(rows, start=2):
            style = row[style_at]
            if style is None or not str(style).strip():
                continue
            # The trailing subtotal row is not a style; see read_sales.
            if not _is_row_number(row[row_at]):
                continue
            index[_norm(style)] = len(kept)
            kept.append(list(row))
            sheet_rows.append(offset)
        return Grid(headers=headers, rows=kept, index=index, sheet_rows=sheet_rows,
                    row_at=row_at)
    finally:
        workbook.close()


def list_sheets(xlsx_path: str | Path) -> list[tuple[str, set[str]]]:
    """Every worksheet in the workbook, with the style numbers it lists.

    Picking the wrong sheet is a silent failure, not an error: a style carries a
    different ``Row#`` in every sheet it appears in, so the figures would come
    out plausible and wrong. The file picker scores each sheet against the style
    numbers actually printed on the line sheet rather than guessing from its
    name, and this is what it scores against.

    Sheets that are not ranked-sales sheets come back with an empty set rather
    than raising — a workbook is allowed to hold working notes.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        found: list[tuple[str, set[str]]] = []
        for name in workbook.sheetnames:
            styles: set[str] = set()
            rows = workbook[name].iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                found.append((name, styles))
                continue
            headers = [_norm(cell) if cell is not None else "" for cell in header]
            columns = {name: index for index, name in enumerate(headers) if name}
            style_at = columns.get(STYLE)
            row_at = _row_num_column(columns, headers)
            if style_at is not None and row_at is not None:
                for row in rows:
                    # Read-only rows can be shorter than the header row.
                    if style_at >= len(row) or row_at >= len(row):
                        continue
                    style = row[style_at]
                    if style is None or not str(style).strip():
                        continue
                    if not _is_row_number(row[row_at]):
                        continue  # the trailing subtotal row; see read_sales
                    styles.add(_norm(style))
            found.append((name, styles))
        return found
    finally:
        workbook.close()


def read_sales(xlsx_path: str | Path, sheet_name: str) -> dict[str, SalesRow]:
    """Map style number to its figures, for one named sheet.

    Raises:
        KeyError: the sheet is missing, or lacks a column the line sheet needs.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"no sheet named {sheet_name!r}; workbook has {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            raise KeyError(f"sheet {sheet_name!r} is empty") from None

        headers = [_norm(cell) if cell is not None else "" for cell in header]
        columns = {name: index for index, name in enumerate(headers) if name}
        row_num_at = _row_num_column(columns, headers)
        missing = [name for name in (STYLE, QTY, SALES) if name not in columns]
        if row_num_at is None:
            missing.insert(0, " / ".join(ROW_NUM_ALIASES[:3]))
        if missing:
            raise KeyError(f"sheet {sheet_name!r} is missing columns: {missing}")

        found: dict[str, SalesRow] = {}
        seen: set[str] = set()
        for row in rows:
            style = row[columns[STYLE]]
            if style is None or not str(style).strip():
                continue
            # Every sheet ends with a subtotal row carrying the literal "SubTtl"
            # in both Row# and Style #. A real style always has an integer Row#.
            if not _is_row_number(row[row_num_at]):
                continue
            style = _norm(style)
            if style in seen:
                # Two rows claiming the same style is ambiguous, and quietly
                # keeping whichever came last writes one of them onto the sheet
                # with no way for anyone to know which. Refuse instead.
                raise KeyError(
                    f"sheet {sheet_name!r} lists {style!r} more than once; "
                    f"the tool cannot tell which row is the real one"
                )
            seen.add(style)
            found[style] = SalesRow(
                style=style,
                row_num=int(row[row_num_at]),
                qty=_whole(row[columns[QTY]]),
                sales=_money(row[columns[SALES]]),
            )
        return found
    finally:
        workbook.close()
